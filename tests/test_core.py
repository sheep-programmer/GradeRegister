"""核心逻辑测试：parser / excel_io / state。运行：python -m unittest tests.test_core -v"""
from __future__ import annotations

import os
import shutil
import tempfile
import unittest
from unittest import mock

from grade_app import parser
from grade_app.excel_io import check_mark_text, load_sheet, save_sheet
from grade_app.state import AppState
from openpyxl import Workbook

CFG = {
    "strip_prefix": True, "strip_suffix": True, "score_cutoff": 0.55,
    "write_formula": True, "write_checked": True, "auto_save": False,
}


# ---------------------------------------------------------------------------
# parser
# ---------------------------------------------------------------------------
class TestCnNumber(unittest.TestCase):
    def test_basic(self):
        cases = {
            "十八": 18, "十二": 12, "二十": 20, "二十五": 25,
            "一百零五": 105, "一十八": 18, "两": 2, "零": 0,
            "十": 10, "三十": 30, "九十九": 99, "一百": 100,
        }
        for s, want in cases.items():
            self.assertEqual(parser.cn_int_to_value(s), want, s)

    def test_float(self):
        self.assertEqual(parser.cn_number_to_float("十二点五"), 12.5)
        self.assertEqual(parser.cn_number_to_float("一点五"), 1.5)
        self.assertEqual(parser.any_number_to_float("18.5"), 18.5)
        self.assertEqual(parser.any_number_to_float("十八"), 18)

    def test_invalid(self):
        self.assertIsNone(parser.cn_int_to_value("abc"))
        self.assertIsNone(parser.cn_number_to_float("十三点"))


class TestExtractScores(unittest.TestCase):
    def test_simple(self):
        self.assertEqual(parser.extract_scores("十八分"), [18.0])
        self.assertEqual(parser.extract_scores("18分"), [18.0])
        self.assertEqual(parser.extract_scores("12"), [12.0])
        self.assertEqual(parser.extract_scores("18.5"), [18.5])

    def test_sequence(self):
        got = parser.extract_scores("十八分十二分十五分十七分")
        self.assertEqual(got, [18.0, 12.0, 15.0, 17.0])
        got2 = parser.extract_scores("18分、12分、15分、17分")
        self.assertEqual(got2, [18.0, 12.0, 15.0, 17.0])

    def test_prefix_stripped(self):
        self.assertEqual(parser.extract_scores("第一题十八分"), [18.0])
        self.assertEqual(parser.extract_scores("第一题18分"), [18.0])
        self.assertEqual(parser.extract_scores("第1题18.5"), [18.5])
        self.assertEqual(parser.extract_scores("第六题六分"), [6.0])
        # 关闭剥离时，题号也会被当成分数
        self.assertEqual(parser.extract_scores("第一题十八分", strip_prefix=False),
                         [1.0, 18.0])

    def test_suffix_off(self):
        self.assertEqual(parser.extract_scores("第一题十八分", strip_suffix=False),
                         [18.0])  # 前缀剥离后"分"字保留但无数字，不影响
        self.assertEqual(parser.extract_scores("十八分", strip_suffix=False), [18.0])

    def test_mixed_text(self):
        got = parser.extract_scores("第一题十八分 第二题十二分")
        self.assertEqual(got, [18.0, 12.0])

    def test_mishear_digit_homophones(self):
        """同音字纠正：失→十、期→七、吧→八，听岔的分数仍能对上。"""
        self.assertEqual(parser.extract_scores("失误"), [15.0])
        self.assertEqual(parser.extract_scores("时期"), [17.0])
        self.assertEqual(parser.extract_scores("是吧"), [18.0])
        self.assertEqual(parser.extract_scores("第漆题三分"), [3.0])


class TestExtractScoreItems(unittest.TestCase):
    def test_single_question(self):
        self.assertEqual(parser.extract_score_items("第三题九分"), [(3, 9.0)])

    def test_no_question_number(self):
        self.assertEqual(parser.extract_score_items("十八分"), [(None, 18.0)])

    def test_question_head_alone_is_placeholder(self):
        self.assertEqual(parser.extract_score_items("第三题"), [(3, None)])

    def test_all_questions_in_one_breath(self):
        """一口气念完四道题：每道题各自的题号都要认出来，不能全归到第一题。"""
        got = parser.extract_score_items(
            "第一题十分第二题二十分第三题三十分第四题四十分")
        self.assertEqual(got, [(1, 10.0), (2, 20.0), (3, 30.0), (4, 40.0)])

    def test_multiple_heads_with_spaces(self):
        got = parser.extract_score_items("第一题十八分 第二题十二分")
        self.assertEqual(got, [(1, 18.0), (2, 12.0)])

    def test_head_after_leading_scores(self):
        """题号前先念了一个不带题号的分数，两者都要保留。"""
        got = parser.extract_score_items("十八分第二题十二分")
        self.assertEqual(got, [(None, 18.0), (2, 12.0)])

    def test_separator_words_split_scores(self):
        self.assertEqual(parser.extract_score_items("10加5加5加3"),
                         [(None, 10.0), (None, 5.0), (None, 5.0), (None, 3.0)])

    def test_one_fen_settles_one_score(self):
        """「第一题十五分」被听成「七月十五分」时只认紧挨着「分」的 15。"""
        self.assertEqual(parser.extract_score_items("七月十五分"), [(None, 15.0)])

    def test_bare_number_without_fen_ignored(self):
        self.assertEqual(parser.extract_score_items("十八"), [])

    def test_inner_head_needs_the_di_character(self):
        """句中的题号必须带「第」字，否则「再说一次十分」里的「一次」会被当题号。"""
        self.assertEqual(parser.extract_score_items("再说一次十分"),
                         [(None, 10.0)])

    def test_leading_head_without_di_still_works(self):
        """句首题号识别时常丢掉「第」字，这里仍要认出来。"""
        self.assertEqual(parser.extract_score_items("三题九分"), [(3, 9.0)])


class TestHeaderScoreItems(unittest.TestCase):
    """按表头名念分数：表头不一定是「第X题」，也可能是「程序题」。"""

    HEADERS = ["程序题", "选择题", "填空题"]
    NUMBERED = ["第一题", "第二题", "第三题", "第四题"]

    def test_exact_header_name(self):
        self.assertEqual(
            parser.extract_score_items("程序题72分", headers=self.HEADERS),
            [(1, 72.0)])

    def test_later_header_name(self):
        self.assertEqual(
            parser.extract_score_items("选择题85分", headers=self.HEADERS),
            [(2, 85.0)])

    def test_several_headers_in_one_breath(self):
        got = parser.extract_score_items("程序题72分选择题85分填空题60分",
                                         headers=self.HEADERS)
        self.assertEqual(got, [(1, 72.0), (2, 85.0), (3, 60.0)])

    def test_question_number_still_works_with_named_headers(self):
        """念序号照旧生效，按位置命中第一个分数列。"""
        self.assertEqual(
            parser.extract_score_items("第一题78分", headers=self.HEADERS),
            [(1, 78.0)])

    def test_numbered_headers_go_through_index_path(self):
        self.assertEqual(
            parser.extract_score_items("第二题二十分", headers=self.NUMBERED),
            [(2, 20.0)])

    def test_header_alone_is_placeholder(self):
        self.assertEqual(
            parser.extract_score_items("程序题", headers=self.HEADERS),
            [(1, None)])

    def test_no_headers_argument_keeps_old_behaviour(self):
        self.assertEqual(parser.extract_score_items("第三题九分"), [(3, 9.0)])

    def test_header_survives_mishear_mapping(self):
        """表头里的字也在同音映射里（实→十）时，匹配必须用原文。"""
        self.assertEqual(
            parser.extract_score_items("实验题七十二分", headers=["实验题"]),
            [(1, 72.0)])

    def test_physics_header_not_damaged_by_mishear_map(self):
        """「误→五」这类映射只作用于数字提取，不能把表头改坏。"""
        self.assertEqual(
            parser.extract_score_items("物理七十二分", headers=["物理"]),
            [(1, 72.0)])


class TestHeaderPinyinCorrection(unittest.TestCase):
    def test_misheard_header_corrected(self):
        text, fixed = parser.correct_headers_in_text("成序题七十二分",
                                                     ["程序题", "选择题"])
        self.assertIn("程序题", text)
        self.assertTrue(fixed)

    def test_numbered_headers_never_pinyin_matched(self):
        """「第一题」「第二题」发音太近，模糊匹配会把分数填到错的题上。"""
        text, fixed = parser.correct_headers_in_text("第二题二十分",
                                                     ["第一题", "第二题"])
        self.assertEqual(fixed, [])
        self.assertEqual(text, "第二题二十分")

    def test_student_name_not_mistaken_for_header(self):
        _text, fixed = parser.correct_headers_in_text(
            "刘洋", ["程序题", "选择题"], names=["刘洋"])
        self.assertEqual(fixed, [])

    def test_exact_header_left_alone(self):
        text, fixed = parser.correct_headers_in_text("程序题72分", ["程序题"])
        self.assertIn("程序题", text)
        self.assertEqual(fixed, [])

    def test_corrected_header_feeds_score_items(self):
        """纠正后的文本要能被分数提取认出来，端到端连上。"""
        headers = ["程序题", "选择题"]
        text, _fixed = parser.correct_headers_in_text("成序题七十二分", headers)
        self.assertEqual(parser.extract_score_items(text, headers=headers),
                         [(1, 72.0)])


class TestStudentIdExtraction(unittest.TestCase):
    """号码必须带标记词，裸数字一律当分数。"""

    def test_plain_number_with_hao(self):
        num, rest = parser.extract_student_id("5号")
        self.assertEqual(parser.any_number_to_float(num), 5)
        self.assertEqual(rest.strip(), "")

    def test_chinese_number_with_hao(self):
        num, _rest = parser.extract_student_id("五号")
        self.assertEqual(parser.any_number_to_float(num), 5)

    def test_xuehao_prefix(self):
        for text in ("学号5", "座号5", "编号五"):
            num, _rest = parser.extract_student_id(text)
            self.assertEqual(parser.any_number_to_float(num), 5, text)

    def test_nth_form(self):
        num, _rest = parser.extract_student_id("第五个")
        self.assertEqual(parser.any_number_to_float(num), 5)

    def test_id_with_scores_keeps_rest(self):
        num, rest = parser.extract_student_id("5号78分85分")
        self.assertEqual(parser.any_number_to_float(num), 5)
        self.assertEqual(parser.extract_scores(rest), [78.0, 85.0])

    def test_score_is_not_an_id(self):
        self.assertIsNone(parser.extract_student_id("5分")[0])

    def test_question_head_is_not_an_id(self):
        self.assertIsNone(parser.extract_student_id("第三题")[0])
        self.assertIsNone(parser.extract_student_id("第三题九分")[0])

    def test_bare_number_is_not_an_id(self):
        self.assertIsNone(parser.extract_student_id("78")[0])

    def test_name_is_not_an_id(self):
        self.assertIsNone(parser.extract_student_id("刘洋")[0])


class TestCollapseRepeats(unittest.TestCase):
    def test_whole_text_repeat_collapsed(self):
        self.assertEqual(parser.collapse_repeats("李四李四李四"), "李四")
        self.assertEqual(parser.collapse_repeats("张三张三"), "张三")

    def test_different_names_left_alone(self):
        self.assertEqual(parser.collapse_repeats("张三张三丰"), "张三张三丰")

    def test_non_repeat_left_alone(self):
        for t in ("刘洋", "第一题十分", "", "程序题72分"):
            self.assertEqual(parser.collapse_repeats(t), t)

    def test_single_char_repeat_left_alone(self):
        """「三三」可能是名字的一部分，单字重复不折叠。"""
        self.assertEqual(parser.collapse_repeats("三三"), "三三")


class TestAmbiguousNameCorrection(unittest.TestCase):
    """班上同时有刘洋和刘阳时，别闷头猜一个。"""

    BOTH = ["刘洋", "刘阳", "王小明"]
    ONLY_ONE = ["刘洋", "王小明"]

    def test_ambiguous_pair_not_rewritten(self):
        text, fixed = parser.correct_names_in_text("刘扬", self.BOTH)
        self.assertEqual(fixed, [])
        self.assertEqual(text, "刘扬")

    def test_clear_winner_still_rewritten(self):
        text, fixed = parser.correct_names_in_text("刘扬", self.ONLY_ONE)
        self.assertEqual(text, "刘洋")
        self.assertTrue(fixed)

    def test_ambiguous_text_yields_multiple_candidates(self):
        """不改写之后，模糊匹配要能给出两个候选供弹窗选择。"""
        got = parser.match_student_names("刘扬", self.BOTH)
        self.assertIn("刘洋", got)
        self.assertIn("刘阳", got)

    def test_prefer_set_breaks_the_tie(self):
        """分数接近时优先还没录完的学生。"""
        best, _r = parser.best_name_by_pinyin("刘扬", self.BOTH,
                                              prefer={"刘阳"})
        self.assertEqual(best, "刘阳")

    def test_prefer_does_not_override_clear_winner(self):
        """差距明显时该谁是谁，回头改分照样点得到人。"""
        best, _r = parser.best_name_by_pinyin(
            "王小明", self.BOTH, prefer={"刘阳"})
        self.assertEqual(best, "王小明")


class TestMatchNames(unittest.TestCase):
    NAMES = ["张三", "李四", "王五", "张三丰", "赵六"]

    def test_exact(self):
        self.assertEqual(parser.match_student_names("张三", self.NAMES), ["张三"])

    def test_contained(self):
        self.assertEqual(parser.match_student_names("张三。", self.NAMES), ["张三"])
        self.assertEqual(parser.match_student_names("张三丰", self.NAMES), ["张三丰"])

    def test_fuzzy(self):
        got = parser.match_student_names("章三", self.NAMES)
        self.assertIn("张三", got)

    def test_find_rows_duplicate(self):
        students = [(1, "张三"), (2, "李四"), (4, "张三")]
        rows = parser.find_student_rows("张三", students)
        self.assertEqual(len(rows), 2)


# ---------------------------------------------------------------------------
# excel_io
# ---------------------------------------------------------------------------
class TestExcel(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "总分"])
        ws.append(["张三", "", "", "", ""])
        ws.append(["李四", "", "", "", ""])
        ws.append(["张三", "", "", "", ""])
        wb.save(self.path)

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_load(self):
        m = load_sheet(self.path, CFG, backup=False)
        self.assertEqual(m.name_col, 0)
        self.assertEqual(m.score_cols, [1, 2, 3])
        self.assertEqual(m.total_col, 4)
        self.assertEqual(m.check_col, 5)
        self.assertEqual(len(m.students), 3)
        self.assertEqual([s.name for s in m.students], ["张三", "李四", "张三"])

    def test_save_formula_and_check(self):
        m = load_sheet(self.path, CFG, backup=False)
        s0 = m.students[0]
        s0.scores = {1: 18.0, 2: 12.0, 3: 15.0}
        s0.total = m.calc_total(s0)
        s0.checked = False  # 不一致 -> 标红
        s1 = m.students[1]
        s1.scores = {1: 20.0, 2: 20.0, 3: 20.0}
        s1.total = m.calc_total(s1)
        s1.checked = True   # 一致 -> ✓
        save_sheet(m, CFG, write_formula=True)

        # 读回验证：分数与核对标记
        m2 = load_sheet(self.path, CFG, backup=False)
        self.assertEqual(m2.students[0].scores, {1: 18.0, 2: 12.0, 3: 15.0})
        self.assertEqual(m2.students[1].scores, {1: 20.0, 2: 20.0, 3: 20.0})
        # 核对结果随表格一起还原，重开表不会丢
        self.assertIs(m2.students[0].checked, False)
        self.assertIs(m2.students[1].checked, True)
        self.assertIs(m2.students[2].checked, None)   # 没核对过的行保持未核对
        # 另一个文件验证：总分写数值 + 核对列内容
        path2 = os.path.join(self.tmp, "num.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "总分"])
        ws.append(["张三", "", "", "", ""])
        wb.save(path2)
        m3 = load_sheet(path2, CFG, backup=False)
        s = m3.students[0]
        s.scores = {1: 18.0, 2: 12.0, 3: 15.0}
        s.total = m3.calc_total(s)   # 45
        s.checked = False
        save_sheet(m3, CFG, write_formula=False)
        from openpyxl import load_workbook as lw
        ws2 = lw(path2).active
        self.assertEqual(ws2.cell(row=2, column=5).value, 45.0)   # 总分数值
        self.assertEqual(ws2.cell(row=2, column=6).value, "不一致")  # 核对列
        self.assertIsNotNone(ws2.cell(row=2, column=6).fill.start_color.rgb)

    def test_blank_rows_keep_empty_total(self):
        """一分没录的学生不写总分：否则 =SUM(空白) 会把全班总分刷成 0。"""
        from openpyxl import load_workbook as lw
        m = load_sheet(self.path, CFG, backup=False)
        s0 = m.students[0]
        s0.scores = {1: 18.0}
        s0.total = m.calc_total(s0)
        save_sheet(m, CFG, write_formula=True)

        ws = lw(self.path).active
        self.assertEqual(ws.cell(row=2, column=5).value, "=SUM(B2:D2)")
        self.assertIsNone(ws.cell(row=3, column=5).value)   # 李四没录分
        self.assertIsNone(ws.cell(row=4, column=5).value)

    def test_manual_check_note_preserved(self):
        """老师自己在核对列写的批注不属于本程序的标记，保存时不能被抹掉。"""
        from openpyxl import load_workbook as lw
        wb = lw(self.path)
        wb.active.cell(row=2, column=6, value="家长已签字")
        wb.save(self.path)

        m = load_sheet(self.path, CFG, backup=False)
        self.assertIsNone(m.students[0].checked)     # 不是可识别的标记
        save_sheet(m, CFG, write_formula=True)
        self.assertEqual(lw(self.path).active.cell(row=2, column=6).value,
                         "家长已签字")

    def test_existing_total_preserved_without_scores(self):
        """老师手填过总分但没填分题得分的行，保存时不能被清掉。"""
        from openpyxl import load_workbook as lw
        wb = lw(self.path)
        wb.active.cell(row=3, column=5, value=88)     # 李四只有总分
        wb.save(self.path)

        m = load_sheet(self.path, CFG, backup=False)
        self.assertEqual(m.students[1].total, 88.0)
        self.assertEqual(m.students[1].scores, {})
        save_sheet(m, CFG, write_formula=True)
        self.assertEqual(lw(self.path).active.cell(row=3, column=5).value, 88)

    def test_header_covers_check_column(self):
        """原表没有核对列时，内存表头要补齐，否则按列索引访问会越界。"""
        m = load_sheet(self.path, CFG, backup=False)
        self.assertGreater(len(m.header), m.check_col)
        self.assertEqual(m.header[m.check_col], "核对")

    def test_existing_check_header_not_renamed(self):
        path = os.path.join(self.tmp, "hascheck.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "总分", "复核"])
        ws.append(["张三", "", "", ""])
        wb.save(path)

        m = load_sheet(path, CFG, backup=False)
        self.assertEqual(m.check_col, 3)
        self.assertEqual(m.header[3], "复核")

    def test_save_without_score_columns(self):
        """姓名列紧邻总分列（识别不出题号列）时保存不能崩。"""
        path = os.path.join(self.tmp, "noscore.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "总分"])
        ws.append(["张三", ""])
        wb.save(path)

        m = load_sheet(path, CFG, backup=False)
        self.assertEqual(m.score_cols, [])
        save_sheet(m, CFG, write_formula=True)       # 不抛 IndexError

    def test_removed_score_cleared_from_file(self):
        """内存里删掉的分数要同步从表格抹掉，否则 =SUM() 还在算旧值。"""
        from openpyxl import load_workbook as lw
        m = load_sheet(self.path, CFG, backup=False)
        s0 = m.students[0]
        s0.scores = {1: 18.0, 2: 12.0}
        s0.total = m.calc_total(s0)
        save_sheet(m, CFG, write_formula=True)
        self.assertEqual(lw(self.path).active.cell(row=2, column=3).value, 12)

        del s0.scores[2]                       # 撤销第二题
        s0.total = m.calc_total(s0)
        save_sheet(m, CFG, write_formula=True)
        ws = lw(self.path).active
        self.assertIsNone(ws.cell(row=2, column=3).value)
        self.assertEqual(ws.cell(row=2, column=2).value, 18)

    def test_all_scores_removed_clears_total(self):
        """分数被清光后总分格也要清掉，留着 =SUM() 会算成 0。"""
        from openpyxl import load_workbook as lw
        m = load_sheet(self.path, CFG, backup=False)
        s0 = m.students[0]
        s0.scores = {1: 18.0}
        s0.total = m.calc_total(s0)
        save_sheet(m, CFG, write_formula=True)

        s0.scores.clear()
        s0.total = m.calc_total(s0)
        save_sheet(m, CFG, write_formula=True)
        ws = lw(self.path).active
        self.assertIsNone(ws.cell(row=2, column=2).value)
        self.assertIsNone(ws.cell(row=2, column=5).value)

    def test_text_score_cell_preserved(self):
        """老师手写的「缺考」这类文字不算分数，保存时不能被当成残留值清掉。"""
        from openpyxl import load_workbook as lw
        wb = lw(self.path)
        wb.active.cell(row=2, column=2, value="缺考")
        wb.save(self.path)

        m = load_sheet(self.path, CFG, backup=False)
        self.assertEqual(m.students[0].scores, {})
        save_sheet(m, CFG, write_formula=True)
        self.assertEqual(lw(self.path).active.cell(row=2, column=2).value, "缺考")

    def test_backup_created(self):
        from grade_app.excel_io import _make_backup
        dst = _make_backup(self.path)
        self.assertTrue(os.path.exists(dst))
        self.assertTrue(os.path.basename(dst).startswith("test-"))

    def test_backup_pruned_keeps_newest(self):
        """备份数量有上限，且留下的是时间戳最新的几个。"""
        from grade_app.excel_io import _prune_backups
        backup_dir = os.path.join(self.tmp, "backups")
        os.makedirs(backup_dir)
        stamps = [f"2026082{d}-100000" for d in range(6)]
        for ts in stamps:
            open(os.path.join(backup_dir, f"test-{ts}.xlsx"), "w").close()
        open(os.path.join(backup_dir, "别的表-20260101-100000.xlsx"), "w").close()

        _prune_backups(backup_dir, "test", ".xlsx", keep=3)

        kept = sorted(os.listdir(backup_dir))
        self.assertEqual([n for n in kept if n.startswith("test-")],
                         [f"test-{ts}.xlsx" for ts in stamps[-3:]])
        self.assertIn("别的表-20260101-100000.xlsx", kept)   # 其他表的备份不受影响


class TestReadChecked(unittest.TestCase):
    def test_marks(self):
        from grade_app.excel_io import _read_checked
        self.assertIs(_read_checked("✓"), True)
        self.assertIs(_read_checked(" 一致 "), True)
        self.assertIs(_read_checked("不一致"), False)
        self.assertIs(_read_checked("✗"), False)
        self.assertIsNone(_read_checked(None))
        self.assertIsNone(_read_checked(""))
        self.assertIsNone(_read_checked("待复核"))


# ---------------------------------------------------------------------------
# state
# ---------------------------------------------------------------------------
class TestState(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "总分"])
        ws.append(["张三", "", "", "", ""])
        ws.append(["李四", "", "", "", ""])
        ws.append(["张三", "", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_duplicate_ask_choice(self):
        r = self.state.select_student("张三")
        self.assertIsNotNone(r.select_choices)
        self.assertEqual(len(r.select_choices), 2)
        # 选择第 4 行（第二个张三）
        r2 = self.state.activate_row(3)
        self.assertEqual(self.state.current.name, "张三")
        self.assertEqual(self.state.current.row, 3)
        self.assertEqual(self.state.phase, "scoring")

    def test_scores_flow(self):
        self.state.select_student("李四")
        r = self.state.add_scores("第一题十八分 第二题十二分 第三题十五分")
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0, 3: 15.0})
        self.assertEqual(self.state.current.total, 45.0)
        # 再补一个分数（已满 -> 温和提示，不覆盖已有分数）
        r2 = self.state.add_scores("第四题五分")
        self.assertTrue(r2.ok)
        self.assertIn("忽略", r2.message)
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0, 3: 15.0})
        # 各题已满、又不带题号 -> 不猜位置，提示补念题号
        r3 = self.state.add_scores("五分")
        self.assertFalse(r3.ok)
        self.assertIn("要改分请带上题号", r3.message)
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0, 3: 15.0})

    def test_scores_then_next_student(self):
        self.state.select_student("张三")
        self.state.activate_row(1)
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        # 念下一个学生名字（无数字）-> 切到新学生
        r = self.state.add_scores("王五")
        self.assertNotIn("王五", [s.name for s in self.state.model.students])
        # 用李四验证切换
        r = self.state.add_scores("李四")
        self.assertEqual(self.state.current.name, "李四")
        self.assertEqual(self.state.phase, "scoring")

    def test_total_check(self):
        self.state.select_student("张三")
        self.state.activate_row(1)
        self.state.add_scores("第一题18分 第二题12分 第三题15分")   # 总分 45
        r = self.state.check_total("总分45")
        self.assertTrue(r.ok)
        self.assertIs(self.state.current.checked, True)
        # 不一致的情况
        self.state.select_student("张三")
        self.state.activate_row(3)
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r2 = self.state.check_total("总分46")
        self.assertIs(self.state.current.checked, False)

    def test_undo(self):
        self.state.select_student("李四")
        self.state.add_scores("第一题18分")
        self.state.undo()
        self.assertEqual(self.state.current.scores, {})

    def test_total_keyword_switches_phase(self):
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r = self.state.handle_text("总分 45")
        self.assertIs(self.state.current.checked, True)

    def test_check_reset_after_score_change(self):
        """核对完又改分：旧的核对结论必须作废，否则错分挂着 ✓。"""
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        self.state.check_total("总分45")
        self.assertIs(self.state.current.checked, True)

        self.state.add_scores("第二题20分")          # 改分
        self.assertIsNone(self.state.current.checked)
        self.assertEqual(self.state.current.total, 53.0)

    def test_check_reset_after_undo(self):
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        self.state.check_total("总分45")
        self.state.undo()
        self.assertIsNone(self.state.current.checked)

    def test_last_edit_col_tracked(self):
        """语音填分也要标出刚填的格子，界面据此高亮。"""
        self.state.select_student("李四")
        self.state.add_scores("第二题12分")
        self.assertEqual(self.state._last_edit_col, 2)
        self.state.select_student("张三")            # 切学生时清掉旧高亮
        self.state.activate_row(1)
        self.assertIsNone(self.state._last_edit_col)

    def test_undo_message_uses_header_text(self):
        """提示里说的是表头原文，不是内部列号。"""
        self.state.select_student("李四")
        self.state.add_scores("第三题15分")
        self.assertIn("第三题", self.state.undo().message)

    def test_no_question_score_message_matches_result(self):
        """带题号的分数照填，只有真正没地方放的那个才提示补题号。"""
        self.state.select_student("李四")
        r = self.state.add_scores("第一题18分 第二题12分 第三题18分 20分")
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0, 3: 18.0})
        self.assertIn("已填 3 题", r.message)
        self.assertIn("要改分请带上题号", r.message)
        self.assertNotIn("没带题号已忽略", r.message)

    def test_total_keyword_yigong(self):
        """「一共92分」：92 是总分，「一」不能被当成题分。"""
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题18分")
        r = self.state.handle_text("一共92分")
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0, 3: 18.0})
        self.assertIn("92", r.message)
        self.assertIs(self.state.current.checked, False)   # 48 ≠ 92

    def test_only_one_score_per_fen(self):
        """「第一题十五分」听成「七月十五分」时只认 15，7 不是分数。"""
        self.state.select_student("李四")
        self.state.add_scores("七月十五分")
        self.assertEqual(self.state.current.scores, {1: 15.0})

    def test_question_head_alone_is_remembered(self):
        """只念了「第三题」：记下题号等分数，不能被当成人名「张三」。"""
        self.state.select_student("李四")
        r = self.state.handle_text("第三题")
        self.assertEqual(self.state.current.name, "李四")
        self.assertEqual(self.state.pending_col(), 3)
        self.assertIn("第三题", r.message)

        self.state.handle_text("九分")
        self.assertEqual(self.state.current.scores, {3: 9.0})
        self.assertIsNone(self.state.pending_col())

    def test_pending_question_dropped_on_student_switch(self):
        self.state.select_student("李四")
        self.state.handle_text("第三题")
        self.state.activate_row(1)       # 换到张三
        self.assertIsNone(self.state.pending_col())

    def test_misheard_total_does_not_fill_scores(self):
        """「总分」听成「钟分」时走核对，不能把总分当题分填进空题。"""
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分")
        r = self.state.handle_text("钟分三十分")
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0})
        self.assertIs(self.state.current.checked, True)
        self.assertIn("一致", r.message)

    def test_scores_then_total_in_one_sentence(self):
        """「…十八分 一共 48」：前半填分、后半核对，分数不能被吞掉。"""
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分")
        r = self.state.handle_text("第三题18分 一共 48")
        self.assertEqual(self.state.current.scores, {1: 18.0, 2: 12.0, 3: 18.0})
        self.assertIs(self.state.current.checked, True)
        self.assertIn("一致", r.message)

    def test_next_student_jumps_to_unfinished(self):
        """当前学生录完后，指令切到名单里下一位还没录完的。"""
        self.state.select_student("张三")
        self.state.activate_row(1)
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r = self.state.next_student()
        self.assertTrue(r.ok)
        self.assertEqual(self.state.current.name, "李四")
        self.assertIn("李四", r.message)

    def test_next_student_wraps_around(self):
        """当前这位录完后，下一位未录的在名单开头时也能绕回去找到。"""
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        self.state.select_student("张三")
        self.state.activate_row(3)     # 第二个张三也录完
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r = self.state.next_student()  # 当前 row3，未录的只剩名单开头的 row1
        self.assertTrue(r.ok)
        self.assertEqual(self.state.current.name, "张三")
        self.assertEqual(self.state.current.row, 1)

    def test_next_student_all_done_message(self):
        """全班录完后再念「下一个」：明确提示，不能再切。"""
        for row in (1, 3):
            self.state.select_student("张三")
            self.state.activate_row(row)
            self.state.add_scores("第一题18分 第二题12分 第三题15分")
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r = self.state.next_student()
        self.assertFalse(r.ok)
        self.assertIn("全班都已录完", r.message)
        self.assertIsNotNone(self.state.current)

    def test_next_command_through_handle_text(self):
        """「下一个」走统一输入入口时同样生效，界面回显原句。"""
        self.state.select_student("张三")
        self.state.activate_row(1)
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r = self.state.handle_text("下一个")
        self.assertEqual(self.state.current.name, "李四")
        self.assertEqual(r.heard_text, "下一个")

    def test_auto_next_after_total_check(self):
        """打开 auto_next：核对一致后自动切到下一位未录学生。"""
        st = AppState(cfg={**CFG, "auto_next": True})
        st.load(load_sheet(self.path, CFG, backup=False))
        st.select_student("张三")
        st.activate_row(1)
        st.add_scores("第一题18分 第二题12分 第三题15分")
        r = st.check_total("总分45")
        self.assertTrue(r.ok)
        self.assertIn("已自动切到下一位：李四", r.message)
        self.assertEqual(st.current.name, "李四")

    def test_auto_next_stays_when_all_done(self):
        """全录完后 auto_next 不再乱切，核对完留在当前学生。"""
        st = AppState(cfg={**CFG, "auto_next": True})
        st.load(load_sheet(self.path, CFG, backup=False))
        st.select_student("李四")
        st.add_scores("第一题18分 第二题12分 第三题15分")
        st.select_student("张三")
        st.activate_row(1)
        st.add_scores("第一题18分 第二题12分 第三题15分")
        st.select_student("张三")
        st.activate_row(3)
        st.add_scores("第一题18分 第二题12分 第三题15分")
        r = st.check_total("总分45")
        self.assertIn("一致", r.message)
        self.assertNotIn("已自动切到下一位", r.message)
        self.assertEqual(st.current.name, "张三")
        self.assertEqual(st.current.row, 3)

    def test_auto_next_off_by_default(self):
        """没开 auto_next 时核对完不切人，等老师自己念下一位。"""
        self.state.select_student("李四")
        self.state.add_scores("第一题18分 第二题12分 第三题15分")
        r = self.state.check_total("总分45")
        self.assertNotIn("已自动切到下一位", r.message)
        self.assertEqual(self.state.current.name, "李四")


class TestNamedHeaderState(unittest.TestCase):
    """表头是「程序题」这类名称时，念表头名就能定向填分。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "named.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "程序题", "选择题", "填空题", "总分"])
        ws.append(["张三", "", "", "", ""])
        ws.append(["吴敏", "", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_fill_by_header_name(self):
        self.state.select_student("张三")
        self.state.handle_text("程序题72分")
        self.assertEqual(self.state.current.scores, {1: 72.0})

    def test_fill_several_headers_in_one_breath(self):
        self.state.select_student("张三")
        self.state.handle_text("程序题72分选择题85分填空题60分")
        self.assertEqual(self.state.current.scores,
                         {1: 72.0, 2: 85.0, 3: 60.0})

    def test_question_number_still_fills(self):
        """念序号照旧按位置填。"""
        self.state.select_student("张三")
        self.state.handle_text("第二题85分")
        self.assertEqual(self.state.current.scores, {2: 85.0})

    def test_misheard_header_corrected_and_filled(self):
        self.state.select_student("张三")
        r = self.state.handle_text("成序题七十二分")
        self.assertEqual(self.state.current.scores, {1: 72.0})
        self.assertIn("程序题", r.heard_text or "")

    def test_message_uses_header_text(self):
        self.state.select_student("张三")
        r = self.state.handle_text("程序题")
        self.assertIn("程序题", r.message)


class TestHeardText(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "heard.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["吴敏", "", "", ""])
        ws.append(["刘洋", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_corrected_name_is_reported(self):
        """听成「五米」而名单里是「吴敏」时，界面要显示纠正后的名字。"""
        r = self.state.handle_text("五米")
        self.assertEqual(self.state.current.name, "吴敏")
        self.assertEqual(r.heard_text, "吴敏")

    def test_unmatched_text_keeps_original(self):
        """没有任何名字发音接近时保留原文，否则不知道到底听到了什么。"""
        r = self.state.handle_text("今天天气不错")
        self.assertIn(r.heard_text or "今天天气不错", "今天天气不错")

    def test_exact_name_needs_no_correction(self):
        r = self.state.handle_text("刘洋")
        self.assertEqual(r.heard_text or "刘洋", "刘洋")


class TestUndoGrouping(unittest.TestCase):
    """一次操作算一组：填一句话、清一整列，都只需撤销一次。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "undo.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "第四题", "总分"])
        ws.append(["张三", "", "", "", "", ""])
        ws.append(["李四", "", "", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_one_sentence_undoes_in_one_step(self):
        self.state.select_student("张三")
        self.state.handle_text("第一题十分第二题二十分第三题三十分第四题四十分")
        self.assertEqual(len(self.state.current.scores), 4)
        r = self.state.undo()
        self.assertTrue(r.ok)
        self.assertEqual(self.state.current.scores, {})

    def test_clear_cells_undoes_in_one_step(self):
        m = self.state.model
        self.state.select_student("张三")
        self.state.handle_text("第一题十分第二题二十分")
        targets = [(m.students[0], c) for c in (1, 2)]
        r = self.state.clear_cells(targets)
        self.assertTrue(r.ok)
        self.assertEqual(m.students[0].scores, {})
        self.state.undo()
        self.assertEqual(m.students[0].scores, {1: 10.0, 2: 20.0})

    def test_clear_whole_column_across_students(self):
        m = self.state.model
        for name in ("张三", "李四"):
            self.state.select_student(name)
            self.state.handle_text("第一题十分")
        targets = [(s, 1) for s in m.students]
        self.state.clear_cells(targets)
        self.assertEqual([s.scores.get(1) for s in m.students], [None, None])
        self.state.undo()
        self.assertEqual([s.scores.get(1) for s in m.students], [10.0, 10.0])

    def test_clear_empty_cells_reports_nothing_to_do(self):
        m = self.state.model
        r = self.state.clear_cells([(m.students[0], 1)])
        self.assertFalse(r.ok)

    def test_clearing_resets_total_and_check(self):
        m = self.state.model
        self.state.select_student("张三")
        self.state.handle_text("第一题十分")
        stu = m.students[0]
        stu.checked = True
        self.state.clear_cells([(stu, 1)])
        self.assertIsNone(stu.checked)
        self.assertEqual(stu.total, 0.0)


class TestIdColumn(unittest.TestCase):
    """有学号列就用真学号，没有就用名单序号。"""

    def _build(self, header, rows):
        tmp = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, tmp, ignore_errors=True)
        path = os.path.join(tmp, "t.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for r in rows:
            ws.append(r)
        wb.save(path)
        st = AppState(cfg=CFG)
        st.load(load_sheet(path, CFG, backup=False))
        return st

    def test_id_column_detected(self):
        st = self._build(["学号", "姓名", "第一题", "总分"],
                         [["1001", "张三", "", ""], ["1002", "李四", "", ""]])
        self.assertEqual(st.model.id_col, 0)
        self.assertEqual([s.sid for s in st.model.students], ["1001", "1002"])
        self.assertEqual(st.model.name_col, 1)
        self.assertEqual(st.model.score_cols, [2])

    def test_no_id_column(self):
        st = self._build(["姓名", "第一题", "总分"],
                         [["张三", "", ""], ["李四", "", ""]])
        self.assertEqual(st.model.id_col, -1)

    def test_select_by_real_student_id(self):
        st = self._build(["学号", "姓名", "第一题", "总分"],
                         [["1001", "张三", "", ""], ["1002", "李四", "", ""]])
        r = st.handle_text("1002号")
        self.assertTrue(r.ok, r.message)
        self.assertEqual(st.current.name, "李四")

    def test_select_by_list_position_without_id_column(self):
        st = self._build(["姓名", "第一题", "总分"],
                         [["张三", "", ""], ["李四", "", ""], ["王五", "", ""]])
        st.handle_text("3号")
        self.assertEqual(st.current.name, "王五")

    def test_id_and_scores_in_one_breath(self):
        st = self._build(["姓名", "第一题", "第二题", "总分"],
                         [["张三", "", "", ""], ["李四", "", "", ""]])
        r = st.handle_text("2号78分85分")
        self.assertEqual(st.current.name, "李四")
        self.assertEqual(st.current.scores, {1: 78.0, 2: 85.0})
        self.assertTrue(r.ok, r.message)

    def test_nth_form_selects_by_position(self):
        st = self._build(["姓名", "第一题", "总分"],
                         [["张三", "", ""], ["李四", "", ""], ["王五", "", ""]])
        st.handle_text("第二个")
        self.assertEqual(st.current.name, "李四")

    def test_out_of_range_id_reports_error(self):
        """念了 20 号而班上只有 3 人：明确报错，不去猜名字填错人。"""
        st = self._build(["姓名", "第一题", "总分"],
                         [["张三", "", ""], ["李四", "", ""], ["王五", "", ""]])
        r = st.handle_text("20号")
        self.assertFalse(r.ok)
        self.assertIsNone(st.current)

    def test_score_not_treated_as_id(self):
        st = self._build(["姓名", "第一题", "第二题", "总分"],
                         [["张三", "", "", ""], ["李四", "", "", ""]])
        st.handle_text("张三")
        st.handle_text("5分")
        self.assertEqual(st.current.name, "张三")
        self.assertEqual(st.current.scores, {1: 5.0})


class TestRepeatAndAmbiguityInState(unittest.TestCase):
    def _build(self, names):
        tmp = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, tmp, ignore_errors=True)
        path = os.path.join(tmp, "t.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        for n in names:
            ws.append([n, "", "", ""])
        wb.save(path)
        st = AppState(cfg=CFG)
        st.load(load_sheet(path, CFG, backup=False))
        return st

    def test_repeated_name_collapsed(self):
        st = self._build(["张三", "李四"])
        r = st.handle_text("李四李四李四")
        self.assertTrue(r.ok, r.message)
        self.assertEqual(st.current.name, "李四")

    def test_repeated_misheard_name_still_matches(self):
        """听错又重复三遍时，不折叠的话整串拼音跟谁都不像，会匹配失败。"""
        st = self._build(["吴敏", "张三"])
        r = st.handle_text("五米五米五米")
        self.assertTrue(r.ok, r.message)
        self.assertEqual(st.current.name, "吴敏")
        self.assertEqual(r.heard_text, "吴敏")

    def test_ambiguous_name_asks_instead_of_guessing(self):
        st = self._build(["刘洋", "刘阳", "王小明"])
        r = st.handle_text("刘扬")
        self.assertIsNotNone(r.select_choices)
        names = {n for _row, n in r.select_choices}
        self.assertEqual(names, {"刘洋", "刘阳"})
        self.assertIsNone(st.current)

    def test_unambiguous_name_still_direct(self):
        st = self._build(["刘洋", "王小明"])
        r = st.handle_text("刘扬")
        self.assertIsNone(r.select_choices)
        self.assertEqual(st.current.name, "刘洋")


class TestChoiceOutOfRange(unittest.TestCase):
    """选人时说岔了不能放行——放行会被号码定位当成「名单第 N 位」。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        for name in ("吴敏", "孙丽", "张伟", "陈静"):
            ws.append([name, "", "", ""])
        wb.save(self.path)
        cfg = dict(CFG)
        cfg["auto_save_mode"] = "manual"
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))
        # 「午门儿」谁都不像但发音接近，走「没太听清」那条非模态候选路
        self.st.handle_text("午门儿")
        self.assertTrue(self.st._pending_choices)
        self.n = len(self.st._pending_choices)
        # 越界的说法：比候选数多一个。候选数随名单变，不写死
        self.over = f"第{'一二三四五六'[self.n]}个"

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_out_of_range_choice_does_not_jump_elsewhere(self):
        """候选只有 N 个却说「第 N+1 个」，绝不能跳到名单第 N+1 位。"""
        r = self.st.handle_text(self.over)
        self.assertFalse(r.ok)
        self.assertIsNone(self.st.current)

    def test_out_of_range_choice_keeps_the_candidates(self):
        """候选留着，老师再说一次就能定，不用重念名字。"""
        self.st.handle_text(self.over)
        self.assertTrue(self.st._pending_choices)
        r = self.st.handle_text("第一个")
        self.assertTrue(r.ok)
        self.assertIsNotNone(self.st.current)

    def test_out_of_range_message_lists_the_candidates(self):
        r = self.st.handle_text(self.over)
        self.assertIn(f"只有 {self.n} 个候选", r.message)
        self.assertTrue(r.select_choices)

    def test_a_valid_choice_still_works(self):
        want = self.st._pending_choices[0][1]
        r = self.st.handle_text("第一个")
        self.assertTrue(r.ok)
        self.assertEqual(self.st.current.name, want)


class TestStudentOrdinal(unittest.TestCase):
    """名单序号：念「第几个」与界面显示必须是同一个数。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "总分"])
        for name in ("王小明", "李四", "张伟"):
            ws.append([name, "", ""])
        wb.save(self.path)
        cfg = dict(CFG)
        cfg["auto_save_mode"] = "manual"
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_ordinal_counts_students_not_excel_rows(self):
        """第一位学生在 Excel 第 2 行，序号必须是 1。"""
        first = self.st.model.students[0]
        self.assertEqual(first.row, 1)          # 0 基，即 Excel 第 2 行
        self.assertEqual(self.st.student_no(first), 1)
        self.assertEqual(self.st.student_no(self.st.model.students[2]), 3)

    def test_spoken_ordinal_selects_that_student(self):
        self.st.handle_text("第三个")
        self.assertEqual(self.st.current.name, "张伟")
        self.assertEqual(self.st.student_no(self.st.current), 3)

    def test_message_reports_both_row_and_ordinal(self):
        """行号给人对 Excel，序号给人对「念第几个」。"""
        r = self.st.handle_text("第三个")
        self.assertIn("第3个", r.message)
        self.assertIn(f"第{self.st.current.row + 1}行", r.message)


class TestSaveOnlyWhenChecked(unittest.TestCase):
    """checked 模式：只把核对对得上的结果同步进 Excel。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["李四", "", "", ""])
        wb.save(self.path)
        cfg = dict(CFG)
        cfg["auto_save_mode"] = "checked"
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))
        self.st.select_student("李四")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _saved(self):
        from openpyxl import load_workbook as lw
        ws = lw(self.path).active
        return [ws.cell(row=2, column=c).value for c in (2, 3)]

    def test_is_the_default_mode(self):
        from grade_app.config import DEFAULT_CONFIG
        self.assertEqual(AppState(cfg=dict(DEFAULT_CONFIG)).auto_save_mode(),
                         "checked")

    def test_filling_scores_does_not_write_yet(self):
        self.st.add_scores("第一题18分 第二题12分")
        self.assertEqual(self._saved(), [None, None])

    def test_a_passing_check_writes(self):
        self.st.add_scores("第一题18分 第二题12分")
        self.st.check_total("总分30")
        self.assertEqual(self._saved(), [18, 12])

    def test_a_failing_check_does_not_write(self):
        """念错了先别同步，让老师改完再说。"""
        self.st.add_scores("第一题18分 第二题12分")
        self.st.check_total("总分28")
        self.assertEqual(self._saved(), [None, None])

    def test_score_mode_still_writes_on_every_score(self):
        """老模式不受影响。"""
        self.st.cfg["auto_save_mode"] = "score"
        self.st.add_scores("第一题18分")
        self.assertEqual(self._saved(), [18, None])

    def test_student_mode_still_writes_on_a_failing_check(self):
        """student 模式下核对不一致照旧写盘，只有 checked 模式才拦。"""
        self.st.cfg["auto_save_mode"] = "student"
        self.st.add_scores("第一题18分 第二题12分")
        self.st.check_total("总分28")
        self.assertEqual(self._saved(), [18, 12])


class TestAutosaveFailureIsVisible(unittest.TestCase):
    """自动保存写不进去（Excel 占着文件）必须留下痕迹，不能静默丢分。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["李四", "", "", ""])
        wb.save(self.path)
        cfg = dict(CFG)
        cfg["auto_save_mode"] = "score"
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))
        self.st.select_student("李四")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_no_error_when_saving_works(self):
        self.st.add_scores("第一题18分")
        self.assertIsNone(self.st.save_error)

    def test_locked_file_records_a_readable_reason(self):
        with mock.patch("grade_app.excel_io.save_sheet",
                        side_effect=PermissionError(13, "被占用")):
            self.st.add_scores("第一题18分")
        self.assertIsNotNone(self.st.save_error)
        self.assertIn("Excel", self.st.save_error)

    def test_scores_stay_in_memory_after_a_failed_save(self):
        """写不进去也不能把已填的分数丢掉。"""
        with mock.patch("grade_app.excel_io.save_sheet",
                        side_effect=PermissionError(13, "被占用")):
            self.st.add_scores("第一题18分")
        self.assertEqual(self.st.model.students[0].scores, {1: 18.0})

    def test_error_clears_once_saving_works_again(self):
        with mock.patch("grade_app.excel_io.save_sheet",
                        side_effect=PermissionError(13, "被占用")):
            self.st.add_scores("第一题18分")
        self.st.add_scores("第二题12分")
        self.assertIsNone(self.st.save_error)

    def test_other_errors_are_recorded_too(self):
        with mock.patch("grade_app.excel_io.save_sheet",
                        side_effect=OSError("磁盘满了")):
            self.st.add_scores("第一题18分")
        self.assertIn("磁盘满了", self.st.save_error)


class TestSwitchingSheets(unittest.TestCase):
    """换表比换学生更彻底，逐项残留都会让新表出现莫名其妙的行为。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _make(self, tag, names, qs):
        path = os.path.join(self.tmp, f"{tag}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名"] + [f"第{i}题" for i in range(1, qs + 1)] + ["总分"])
        for n in names:
            ws.append([n] + [None] * (qs + 1))
        wb.save(path)
        return path

    def _state_on_first_sheet(self):
        cfg = dict(CFG, auto_save_mode="manual")
        st = AppState(cfg=cfg)
        st.load(load_sheet(self._make("a", ("张三", "李四"), 2), cfg,
                           backup=False))
        return st

    def _switch(self, st):
        st.load(load_sheet(self._make("b", ("王五", "赵六"), 3), st.cfg,
                           backup=False))

    def test_pending_question_does_not_cross_sheets(self):
        """在 A 表念了「第一题」还没念分数就换表，这个题号不能带过去。

        带过去的话，新表里念的第一个分数会被填到第一题，而老师以为是
        按顺序填下一个空题。
        """
        st = self._state_on_first_sheet()
        st.handle_text("张三")
        st.handle_text("第一题")
        self.assertIsNotNone(st._pending_q)
        self._switch(st)
        self.assertIsNone(st._pending_q)

    def test_last_edited_cell_does_not_cross_sheets(self):
        """旧的编辑列会在新表上亮起一个从没填过的格子。"""
        st = self._state_on_first_sheet()
        st.handle_text("张三")
        st.handle_text("第一题10分")
        self.assertIsNotNone(st._last_edit_col)
        self._switch(st)
        self.assertIsNone(st._last_edit_col)

    def test_held_line_does_not_cross_sheets(self):
        """待补整句指的是旧表里的学生，换表必须作废。"""
        st = self._state_on_first_sheet()
        st._pending_line = "张三第一题10分总分10"
        self._switch(st)
        self.assertIsNone(st._pending_line)

    def test_candidates_and_undo_do_not_cross_sheets(self):
        st = self._state_on_first_sheet()
        st.handle_text("张三")
        st.handle_text("第一题10分")
        st._pending_choices = [(1, "张三")]
        self._switch(st)
        self.assertIsNone(st._pending_choices)
        self.assertEqual(st.undo_stack, [])
        self.assertIsNone(st.current)
        self.assertEqual(st.phase, "idle")

    def test_save_error_does_not_cross_sheets(self):
        """上一份表写不进去的原因，不该挂在新表上报警。"""
        st = self._state_on_first_sheet()
        st.save_error = "自动保存失败：表格正被 Excel 占用"
        self._switch(st)
        self.assertIsNone(st.save_error)

    def test_old_roster_is_gone_after_switching(self):
        st = self._state_on_first_sheet()
        st.handle_text("张三")
        self._switch(st)
        r = st.handle_text("张三")
        self.assertFalse(r.ok)
        self.assertIsNone(st.current)

    def test_switching_students_keeps_the_held_line(self):
        """待补整句是跨激活传递的，切学生时不能被当成临时状态清掉。"""
        st = self._state_on_first_sheet()
        st._pending_line = "李四第一题10分总分10"
        st._activate("张三", st.model.students[0].row, exact=True)
        self.assertIsNotNone(st._pending_line)


class TestUndoStack(unittest.TestCase):
    """撤销栈的极端：空栈、深栈、跨学生、整批清空。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "t.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "总分"])
        for n in ("张三", "李四", "王五"):
            ws.append([n, None, None, None, None])
        wb.save(self.path)
        cfg = dict(CFG, auto_save_mode="manual")
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_undo_on_an_empty_stack_is_harmless(self):
        for _ in range(50):
            r = self.st.undo()
            self.assertFalse(r.ok)

    def test_deep_stack_unwinds_completely(self):
        self.st.handle_text("张三")
        for i in range(300):
            self.st.handle_text(f"第一题{i % 100}分")
        depth = len(self.st.undo_stack)
        for _ in range(depth + 20):
            self.st.undo()
        self.assertEqual(self.st.model.students[0].scores, {})
        self.assertEqual(self.st.undo_stack, [])

    def test_undo_only_touches_the_last_change(self):
        self.st.handle_text("张三")
        self.st.handle_text("第一题10分")
        self.st.handle_text("李四")
        self.st.handle_text("第一题20分")
        self.st.undo()
        self.assertEqual(self.st.model.students[0].scores, {1: 10.0})
        self.assertEqual(self.st.model.students[1].scores, {})

    def test_undo_voids_the_check_result(self):
        self.st.handle_text("张三")
        self.st.handle_text("第一题10分第二题20分第三题30分")
        self.st.handle_text("总分60")
        self.assertIs(self.st.model.students[0].checked, True)
        self.st.undo()
        self.assertIsNone(self.st.model.students[0].checked)

    def test_clearing_everything_is_one_undo(self):
        for n in ("张三", "李四", "王五"):
            self.st.handle_text(n)
            self.st.handle_text("第一题1分第二题2分第三题3分")
        before = [dict(s.scores) for s in self.st.model.students]
        self.st.clear_cells([(s, c) for s in self.st.model.students
                             for c in self.st.model.score_cols])
        self.st.undo()
        self.assertEqual([dict(s.scores) for s in self.st.model.students],
                         before)


class TestMessySheets(unittest.TestCase):
    """老师的真实表格常常不规整，结构识别不能想当然。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _sheet(self, rows):
        wb = Workbook()
        ws = wb.active
        for r in rows:
            ws.append(r)
        wb.save(self.path)
        return load_sheet(self.path, dict(CFG), backup=False)

    def test_header_below_a_title_row(self):
        """表格上方有标题行时，别把标题当表头、把真表头当学生。"""
        m = self._sheet([["三年级期中成绩"], [],
                         ["姓名", "第一题", "第二题", "总分"],
                         ["张三", None, None, None],
                         ["李四", None, None, None]])
        self.assertEqual(m.header_row, 3)
        self.assertEqual(m.header[:4], ["姓名", "第一题", "第二题", "总分"])
        self.assertEqual([s.name for s in m.students], ["张三", "李四"])
        self.assertEqual(m.score_cols, [1, 2])

    def test_title_row_survives_a_save(self):
        """写回时不能把标题行覆盖掉，行号也要对得上。"""
        from openpyxl import load_workbook as lw
        m = self._sheet([["三年级期中成绩"], [],
                         ["姓名", "第一题", "总分"],
                         ["张三", None, None]])
        st = AppState(cfg=dict(CFG, auto_save_mode="manual"))
        st.load(m)
        st.handle_text("张三")
        st.handle_text("第一题18分")
        save_sheet(m, st.cfg, write_formula=True)
        ws = lw(self.path).active
        self.assertEqual(ws.cell(row=1, column=1).value, "三年级期中成绩")
        self.assertEqual(ws.cell(row=3, column=1).value, "姓名")
        self.assertEqual(ws.cell(row=4, column=2).value, 18)

    def test_total_column_in_the_middle(self):
        """总分不在最右边时，题目列全在它右边，也得认出来。"""
        m = self._sheet([["姓名", "总分", "第一题", "第二题"],
                         ["张三", None, None, None]])
        self.assertEqual(m.score_cols, [2, 3])
        self.assertEqual(m.score_count, 2)

    def test_can_actually_score_with_total_in_the_middle(self):
        m = self._sheet([["姓名", "总分", "第一题", "第二题"],
                         ["张三", None, None, None]])
        st = AppState(cfg=dict(CFG, auto_save_mode="manual"))
        st.load(m)
        st.handle_text("张三")
        st.handle_text("第一题10分第二题20分")
        self.assertEqual(m.students[0].scores, {2: 10.0, 3: 20.0})
        self.assertEqual(m.calc_total(m.students[0]), 30.0)

    def test_plain_sheet_still_uses_the_first_row(self):
        """规整的表不受影响。"""
        m = self._sheet([["姓名", "第一题", "总分"], ["张三", None, None]])
        self.assertEqual(m.header_row, 1)
        self.assertEqual([s.name for s in m.students], ["张三"])


class TestAbsurdScores(unittest.TestCase):
    """识别偶尔把一串数字连成天文数字，填进去会毁掉整张表。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["张三", "", "", ""])
        wb.save(self.path)
        cfg = dict(CFG, auto_save_mode="manual")
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))
        self.st.handle_text("张三")
        self.stu = self.st.model.students[0]

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_astronomical_value_is_refused(self):
        r = self.st.add_scores("第一题99999999999999999999分")
        self.assertEqual(self.stu.scores, {})
        self.assertIn("不像分数", r.message)

    def test_the_limit_itself_is_accepted(self):
        from grade_app.state import MAX_SCORE
        self.st.add_scores(f"第一题{MAX_SCORE:.0f}分")
        self.assertEqual(self.stu.scores, {1: MAX_SCORE})

    def test_just_over_the_limit_is_refused(self):
        from grade_app.state import MAX_SCORE
        self.st.add_scores(f"第一题{MAX_SCORE + 1:.0f}分")
        self.assertEqual(self.stu.scores, {})

    def test_a_good_score_in_the_same_breath_still_lands(self):
        """一句里一个正常一个荒谬，正常那个不能被牵连。"""
        self.st.add_scores("第一题10分第二题99999999分")
        self.assertEqual(self.stu.scores, {1: 10.0})

    def test_normal_scores_are_untouched(self):
        self.st.add_scores("第一题100分第二题0分")
        self.assertEqual(self.stu.scores, {1: 100.0, 2: 0.0})


class TestExcelCentring(unittest.TestCase):
    """写回 Excel 时表头与各学生行居中，跟软件里看到的一致。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["张三", "", "", ""])
        wb.save(self.path)
        self.cfg = dict(CFG)

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _saved(self):
        from openpyxl import load_workbook as lw
        return lw(self.path).active

    def _fill_and_save(self):
        m = load_sheet(self.path, self.cfg, backup=False)
        m.students[0].scores = {1: 10.0, 2: 20.0}
        m.students[0].total = 30.0
        save_sheet(m, self.cfg, write_formula=True)
        return m

    def test_header_row_is_centred(self):
        m = self._fill_and_save()
        ws = self._saved()
        for col in range(m.check_col + 1):
            self.assertEqual(ws.cell(row=1, column=col + 1).alignment.horizontal,
                             "center")

    def test_student_rows_are_centred(self):
        m = self._fill_and_save()
        ws = self._saved()
        for col in range(m.check_col + 1):
            cell = ws.cell(row=2, column=col + 1)
            self.assertEqual(cell.alignment.horizontal, "center")
            self.assertEqual(cell.alignment.vertical, "center")

    def test_centring_keeps_the_mismatch_colours(self):
        """居中不能把不一致那格的红底红字冲掉。"""
        m = load_sheet(self.path, self.cfg, backup=False)
        stu = m.students[0]
        stu.scores = {1: 10.0, 2: 20.0}
        stu.total = 30.0
        stu.checked = False
        stu.spoken_total = 28.0
        save_sheet(m, self.cfg, write_formula=True)
        cell = self._saved().cell(row=2, column=m.check_col + 1)
        self.assertEqual(cell.alignment.horizontal, "center")
        self.assertEqual(cell.value, "不一致（报28 算30）")
        self.assertEqual(cell.fill.start_color.rgb, "FFFFC7CE")

    def test_centring_keeps_wrap_text(self):
        """只改水平/垂直对齐，老师设过的换行不能被抹掉。"""
        from openpyxl import load_workbook as lw
        from openpyxl.styles import Alignment
        wb = lw(self.path)
        wb.active.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
        wb.save(self.path)
        self._fill_and_save()
        self.assertTrue(self._saved().cell(row=2, column=1).alignment.wrap_text)

    def test_check_column_widens_for_the_mismatch_text(self):
        """「不一致（报28 算30）」按默认列宽会显示成 ####。"""
        from openpyxl.utils import get_column_letter
        m = load_sheet(self.path, self.cfg, backup=False)
        stu = m.students[0]
        stu.scores = {1: 10.0, 2: 20.0}
        stu.total = 30.0
        stu.checked = True
        save_sheet(m, self.cfg, write_formula=True)
        letter = get_column_letter(m.check_col + 1)
        with_tick = self._saved().column_dimensions[letter].width
        stu.checked = False
        stu.spoken_total = 28.0
        save_sheet(m, self.cfg, write_formula=True)
        with_text = self._saved().column_dimensions[letter].width
        self.assertGreater(with_text, with_tick)
        # 「不一致（报28 算30）」按半角计 20 格，再留 2 格余量
        self.assertGreaterEqual(with_text, 20)

    def test_a_wider_check_column_is_left_alone(self):
        """老师自己拉宽过就别动他的。"""
        from openpyxl import load_workbook as lw
        from openpyxl.utils import get_column_letter
        m = load_sheet(self.path, self.cfg, backup=False)
        letter = get_column_letter(m.check_col + 1)
        wb = lw(self.path)
        wb.active.column_dimensions[letter].width = 40
        wb.save(self.path)
        self._fill_and_save()
        self.assertEqual(self._saved().column_dimensions[letter].width, 40)


class TestWholeLineForAnotherStudent(unittest.TestCase):
    """一句话念完另一位学生的「名字＋各题＋总分」，不能牵连上一位。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        for name in ("张三", "孙丽", "郑浩", "张三"):   # 张三重名两行
            ws.append([name, "", "", ""])
        wb.save(self.path)
        cfg = dict(CFG)
        cfg["auto_save_mode"] = "manual"
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))
        self.zheng = self._stu("郑浩")[0]
        # 郑浩先录完并核对通过
        self.st.activate_row(self.zheng.row)
        self.st.handle_text("第一题十分第二题十分总分二十分")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _stu(self, name):
        return [s for s in self.st.model.students if s.name == name]

    def test_setup_left_the_previous_student_checked(self):
        self.assertIs(self.zheng.checked, True)

    def test_another_students_line_does_not_touch_the_current_one(self):
        """核心：念孙丽那句的总分，绝不能拿去核对还停在当前的郑浩。"""
        self.st.handle_text("孙丽第一题七十分第二题六十分总分一百三十分")
        self.assertIs(self.zheng.checked, True)
        self.assertIsNone(self.zheng.spoken_total)

    def test_another_students_line_lands_on_that_student(self):
        self.st.handle_text("孙丽第一题七十分第二题六十分总分一百三十分")
        sun = self._stu("孙丽")[0]
        self.assertEqual(sun.scores, {1: 70.0, 2: 60.0})
        self.assertIs(sun.checked, True)
        self.assertEqual(self.st.current.name, "孙丽")

    def test_duplicate_name_asks_first_and_spares_the_previous_student(self):
        """同名两位要弹窗；在选定之前，上一位的核对结论不许动。"""
        r = self.st.handle_text("张三第一题十分第二题二十分总分三十分")
        self.assertTrue(r.select_choices)
        self.assertEqual({n for _row, n in r.select_choices}, {"张三"})
        self.assertIs(self.zheng.checked, True)
        for stu in self._stu("张三"):
            self.assertEqual(stu.scores, {})

    def test_picking_the_duplicate_applies_the_held_line(self):
        """选完人，这一句的分数与总分要补到选中的那位身上。"""
        self.st.handle_text("张三第一题十分第二题二十分总分三十分")
        first, second = self._stu("张三")
        self.st.activate_row(first.row)
        self.assertEqual(first.scores, {1: 10.0, 2: 20.0})
        self.assertIs(first.checked, True)
        self.assertEqual(second.scores, {})      # 另一位张三不受影响

    def test_a_wrong_total_still_lands_on_the_right_student(self):
        """总分念错时，标红的必须是这一句点到的人。"""
        self.st.handle_text("孙丽第一题七十分第二题六十分总分一百分")
        sun = self._stu("孙丽")[0]
        self.assertIs(sun.checked, False)
        self.assertEqual(sun.spoken_total, 100.0)
        self.assertIs(self.zheng.checked, True)

    def test_held_line_expires_when_the_teacher_says_something_else(self):
        """没选人就改口，暂存作废；之后随手点一行不该被灌上旧分数。"""
        self.st.handle_text("张三第一题十分第二题二十分总分三十分")
        self.st.handle_text("孙丽")
        first = self._stu("张三")[0]
        self.st.activate_row(first.row)
        self.assertEqual(first.scores, {})

    def test_own_name_repeated_still_fills_and_checks(self):
        """念的是当前学生自己的名字，照旧走「填分＋核对」，没被拦掉。"""
        sun = self._stu("孙丽")[0]
        self.st.activate_row(sun.row)
        self.st.handle_text("孙丽第一题七十分第二题六十分总分一百三十分")
        self.assertEqual(sun.scores, {1: 70.0, 2: 60.0})
        self.assertIs(sun.checked, True)


class TestTotalCheckFeedback(unittest.TestCase):
    """核对总分的反馈：一致与不一致要能分辨，差多少要写出来。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["李四", "", "", ""])
        wb.save(self.path)
        cfg = dict(CFG)
        cfg["auto_save_mode"] = "manual"
        self.st = AppState(cfg=cfg)
        self.st.load(load_sheet(self.path, cfg, backup=False))
        self.st.select_student("李四")
        self.st.add_scores("第一题18分 第二题12分")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_matching_total_asks_for_the_success_sound(self):
        r = self.st.check_total("总分30")
        self.assertEqual(r.sound, "success")
        self.assertTrue(r.ok)

    def test_mismatching_total_asks_for_the_error_sound(self):
        """念错要响报错音，不能跟填对一个分数一个声。"""
        r = self.st.check_total("总分28")
        self.assertEqual(r.sound, "error")
        self.assertFalse(r.ok)

    def test_mismatch_reports_the_gap(self):
        r = self.st.check_total("总分28")
        self.assertIn("28", r.message)
        self.assertIn("30", r.message)
        self.assertIn("-2", r.message)      # 差值直接写出来

    def test_mismatch_keeps_what_the_teacher_said(self):
        """核对列要连老师报的数一起写，只写「不一致」看不出差在哪。"""
        self.st.check_total("总分28")
        stu = self.st.model.students[0]
        self.assertEqual(stu.spoken_total, 28.0)
        self.assertEqual(check_mark_text(stu), "不一致（报28 算30）")

    def test_matching_total_clears_any_earlier_spoken_value(self):
        self.st.check_total("总分28")
        self.st.check_total("总分30")
        self.assertIsNone(self.st.model.students[0].spoken_total)

    def test_changing_a_score_voids_the_spoken_value(self):
        """分数一改，上一轮报的总分就不该再挂在核对列上。"""
        self.st.check_total("总分28")
        self.st.add_scores("第一题20分")
        self.assertIsNone(self.st.model.students[0].spoken_total)
        self.assertIsNone(self.st.model.students[0].checked)

    def test_undo_voids_the_spoken_value(self):
        self.st.check_total("总分28")
        self.st.undo()
        self.assertIsNone(self.st.model.students[0].spoken_total)

    def test_spoken_value_survives_a_reopen(self):
        """存盘关掉再打开，仍然看得见当时报的是多少。"""
        self.st.check_total("总分28")
        save_sheet(self.st.model, self.st.cfg, write_formula=True)
        again = load_sheet(self.path, self.st.cfg, backup=False)
        stu = again.students[0]
        self.assertIs(stu.checked, False)
        self.assertEqual(stu.spoken_total, 28.0)
        self.assertEqual(check_mark_text(stu), "不一致（报28 算30）")

    def test_total_column_survives_a_reopen_with_formulas(self):
        """总分列存的是 =SUM()，读回来是公式字符串，得自己算一遍。"""
        save_sheet(self.st.model, self.st.cfg, write_formula=True)
        again = load_sheet(self.path, self.st.cfg, backup=False)
        self.assertEqual(again.students[0].total, 30.0)


class TestAutoSaveMode(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "test.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["李四", "", "", ""])
        wb.save(self.path)

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _state(self, **over):
        cfg = dict(CFG)
        cfg["auto_save"] = True
        cfg.update(over)
        st = AppState(cfg=cfg)
        st.load(load_sheet(self.path, cfg, backup=False))
        st.select_student("李四")
        return st

    def _saved_scores(self):
        from openpyxl import load_workbook as lw
        ws = lw(self.path).active
        return [ws.cell(row=2, column=c).value for c in (2, 3)]

    def test_mode_score_writes_every_time(self):
        st = self._state(auto_save_mode="score")
        st.add_scores("第一题18分")
        self.assertEqual(self._saved_scores(), [18, None])

    def test_mode_student_waits_for_the_last_question(self):
        st = self._state(auto_save_mode="student")
        st.add_scores("第一题18分")
        self.assertEqual(self._saved_scores(), [None, None])   # 还没录完
        st.add_scores("第二题12分")
        self.assertEqual(self._saved_scores(), [18, 12])

    def test_mode_manual_never_writes(self):
        st = self._state(auto_save_mode="manual")
        st.add_scores("第一题18分 第二题12分")
        st.check_total("总分30")
        self.assertEqual(self._saved_scores(), [None, None])

    def test_trailing_question_head_still_saves_the_score(self):
        """「第一题18分第二题」句尾只有题号：前面那个分数照样要落盘。"""
        st = self._state(auto_save_mode="score")
        st.add_scores("第一题18分第二题")
        self.assertEqual(self._saved_scores(), [18, None])

    def test_legacy_auto_save_flag_maps_to_manual(self):
        st = self._state(auto_save=False, auto_save_mode=None)
        self.assertEqual(st.auto_save_mode(), "manual")
        st.add_scores("第一题18分")
        self.assertEqual(self._saved_scores(), [None, None])


class TestSplitAtTotal(unittest.TestCase):
    def test_markers(self):
        self.assertEqual(parser.split_at_total("第一题18分 一共 48"),
                         ("第一题18分 ", " 48"))
        self.assertEqual(parser.split_at_total("总分45"), ("", "45"))
        self.assertEqual(parser.split_at_total("总成绩 45"), ("", " 45"))
        self.assertEqual(parser.split_at_total("第一题18分"), ("第一题18分", ""))

    def test_misheard_total_by_pinyin(self):
        """「总分」被听成同音词时照样切得出来，否则总分会被当成题分填掉。"""
        for text in ("钟分六十分", "充分六十分", "钟晨六十分", "中分六十分",
                     "钟分 六十"):
            with self.subTest(text=text):
                self.assertTrue(parser.find_total_keyword(text), text)
                _head, tail = parser.split_at_total(text)
                self.assertEqual(parser.extract_scores(tail), [60.0])

    def test_pinyin_guess_needs_a_number(self):
        """没有数字的句子不做同音猜测，避免把姓名当成总分。"""
        self.assertFalse(parser.find_total_keyword("钟分"))
        self.assertFalse(parser.find_total_keyword("充分"))

    def test_pinyin_guess_skips_roster_names(self):
        """班里真有同音姓名时，念名字不能被当成念总分。"""
        roster = ["张三", "钟芳", "宋芬"]
        self.assertFalse(parser.find_total_keyword("宋芬十八分", roster))
        self.assertTrue(parser.find_total_keyword("宋芬十八分", ["张三"]))

    def test_score_sentences_not_mistaken_for_total(self):
        for text in ("第一题十八分", "十八分十二分", "满分", "三分", "得分"):
            with self.subTest(text=text):
                self.assertFalse(parser.find_total_keyword(text), text)


if __name__ == "__main__":
    unittest.main()

class TestNameAndScoresInOneBreath(unittest.TestCase):
    """「赵磊第一题十分第二题十二分」这类连念，在任何阶段都要填得进去。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "t.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "第四题", "总分"])
        for n in ("赵磊", "张三", "李四"):
            ws.append([n, "", "", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _scores(self):
        m = self.state.model
        return [self.state.current.scores.get(c) for c in m.score_cols]

    def test_fills_from_idle_phase(self):
        """刚启动（idle）就连名带分念一整句。"""
        self.assertEqual(self.state.phase, "idle")
        self.state.handle_text("赵磊第一题十分第二题十二分")
        self.assertEqual(self.state.current.name, "赵磊")
        self.assertEqual(self._scores(), [10.0, 12.0, None, None])

    def test_fills_from_scoring_phase(self):
        self.state.handle_text("张三")
        self.state.handle_text("赵磊第一题十分第二题十二分")
        self.assertEqual(self.state.current.name, "赵磊")
        self.assertEqual(self._scores(), [10.0, 12.0, None, None])

    def test_real_world_misheard_sentence(self):
        """实际录到的一句：「是跟」是「十分」的误听，题号超范围的那个忽略掉。"""
        r = self.state.handle_text(
            "赵磊第一题是跟第二题十二分第三题七分第八题七十五分")
        self.assertEqual(self.state.current.name, "赵磊")
        self.assertEqual(self._scores()[:3], [10.0, 12.0, 7.0])
        self.assertIn("超出范围", r.message)

    def test_plain_name_still_only_switches(self):
        """只念名字不能凭空填分。"""
        self.state.handle_text("赵磊")
        self.assertEqual(self.state.current.name, "赵磊")
        self.assertEqual(self._scores(), [None] * 4)

    def test_scores_without_name_need_a_current_student(self):
        """idle 阶段只念分数、又没有当前学生时，要说清楚而不是崩。"""
        r = self.state.handle_text("第一题十分")
        self.assertFalse(r.ok)
        self.assertIsNone(self.state.current)


class TestPinyinSimilarity(unittest.TestCase):
    """按字比声母韵母：整串比会把「村你/孙丽」这种模糊音结构丢掉。"""

    def test_fuzzy_initials_and_finals(self):
        """c↔s、n↔l 是普通话最常混的几组，韵母完全相同时要给高分。"""
        self.assertGreater(parser.pinyin_similarity("村你", "孙丽"), 0.7)

    def test_identical_pronunciation_scores_one(self):
        for a, b in (("正好", "郑浩"), ("例四", "李四"), ("沉静", "陈静")):
            self.assertAlmostEqual(parser.pinyin_similarity(a, b), 1.0,
                                   delta=0.01, msg=f"{a}/{b}")

    def test_different_people_score_low(self):
        """不同的人必须拉开距离，否则会把分数填到别人那行。"""
        for a, b in (("张三", "李四"), ("刘洋", "陈静"),
                     ("王小明", "赵磊"), ("吴敏", "周杰")):
            self.assertLess(parser.pinyin_similarity(a, b), 0.35,
                            msg=f"{a}/{b}")

    def test_length_mismatch_penalised(self):
        """「张三」不能蹭上「张三丰」，否则三个字的名字永远选不中。"""
        self.assertLess(parser.pinyin_similarity("张三", "张三丰"), 0.8)

    def test_handles_empty_and_non_chinese(self):
        for a, b in (("", "张三"), ("abc", "张三"), ("张三", "")):
            v = parser.pinyin_similarity(a, b)
            self.assertGreaterEqual(v, 0.0)
            self.assertLessEqual(v, 1.0)

    def test_threshold_separates_real_cases(self):
        """实测样本：该纠的过阈值，不该纠的不过。"""
        th = parser._NAME_FIX_THRESHOLD
        for got, want in (("村你", "孙丽"), ("正好", "郑浩"), ("五米", "吴敏")):
            self.assertGreaterEqual(parser.pinyin_similarity(got, want), th,
                                    msg=f"{got}->{want} 应该能纠回来")
        for got, want in (("张三", "李四"), ("刘洋", "陈静")):
            self.assertLess(parser.pinyin_similarity(got, want), th,
                            msg=f"{got}->{want} 不该被纠")


class TestScoreTextIsNotAName(unittest.TestCase):
    """纯分数绝不能被当人名——「十三分」和「张三」共享一个「三」字。"""

    NAMES = ["张三", "李四", "王五"]

    def test_score_with_fen_is_number_text(self):
        for t in ("十三分", "13分", "三分", "二十分", "十二分"):
            self.assertTrue(parser.is_number_text(t), t)

    def test_score_does_not_match_any_name(self):
        for t in ("十三分", "三分", "十五分"):
            self.assertEqual(parser.match_student_names(t, self.NAMES), [], t)

    def test_real_names_still_match(self):
        self.assertEqual(parser.match_student_names("张三", self.NAMES), ["张三"])
        self.assertIn("王五", parser.match_student_names("王虎", self.NAMES))

    def test_name_containing_number_char_still_a_name(self):
        """「王五」「李十一」这类名字本身带数字字，不能被当成分数。"""
        names = ["王五", "李十一", "张三"]
        self.assertFalse(parser.is_number_text("王五"))
        self.assertFalse(parser.is_number_text("李十一"))
        self.assertEqual(parser.match_student_names("王五", names), ["王五"])

    def test_question_head_not_a_number_text(self):
        self.assertFalse(parser.is_number_text("第一题"))


class TestScoreTextInState(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "t.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "第三题", "总分"])
        for n in ("张三", "李四", "王五"):
            ws.append([n, "", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_score_does_not_switch_student(self):
        """选中李四后念「十三分」，必须填到李四，不能跳去张三。"""
        self.state.handle_text("李四")
        self.assertEqual(self.state.current.name, "李四")
        self.state.handle_text("十三分")
        self.assertEqual(self.state.current.name, "李四")
        self.assertEqual(self.state.current.scores.get(1), 13.0)

    def test_score_before_any_student_asks_for_a_name(self):
        r = self.state.handle_text("十三分")
        self.assertIsNone(self.state.current)
        self.assertFalse(r.ok)


class TestDuplicateNamesDoNotBlockCorrection(unittest.TestCase):
    """班里有重名时，拼音纠错不能因此瘫掉。

    同名两位的拼音必然一模一样，那是重名、该交给候选框按行选，
    不是「听不清是甲还是乙」——把它当歧义会让整条纠错通道失效。
    """

    DUP = ["王小明", "李四", "张伟", "张三", "孙丽", "张三"]

    def test_corrects_even_with_duplicates(self):
        text, fixed = parser.correct_names_in_text("长三", self.DUP)
        self.assertEqual(text, "张三", f"重名把纠错挡住了：{fixed}")
        self.assertTrue(fixed)

    def test_same_name_twice_is_not_ambiguous(self):
        self.assertFalse(parser.is_name_ambiguous("长三", self.DUP))

    def test_genuinely_close_names_still_ambiguous(self):
        """真正发音相近的两个不同名字，仍然要判为分不清。"""
        self.assertTrue(parser.is_name_ambiguous("刘扬", ["刘洋", "刘阳", "王五"]))

    def test_ranking_lists_each_name_once(self):
        ranked = parser.rank_names_by_pinyin("长三", self.DUP)
        names = [n for n, _ in ranked]
        self.assertEqual(len(names), len(set(names)))

    def test_duplicate_still_offers_row_choice(self):
        """纠对之后，两行同名仍然要让老师选是哪一行。"""
        students = [(i, n) for i, n in enumerate(self.DUP)]
        rows = parser.find_student_rows("张三", students)
        self.assertEqual(len(rows), 2)


class TestMisheardQuestionHead(unittest.TestCase):
    """「题」常被听成「期」，题号丢了整句分数就落到错的列上。"""

    def test_qi_reads_as_ti(self):
        self.assertEqual(parser.extract_score_items("第二期八分"), [(2, 8.0)])
        self.assertEqual(parser.extract_score_items("第五期六分"), [(5, 6.0)])

    def test_several_in_one_breath(self):
        got = parser.extract_score_items("第一期八分第二期七分第三期九分")
        self.assertEqual(got, [(1, 8.0), (2, 7.0), (3, 9.0)])

    def test_normal_ti_unaffected(self):
        self.assertEqual(parser.extract_score_items("第二题八分"), [(2, 8.0)])

    def test_qi_without_number_is_not_a_head(self):
        """「星期」「学期」这类词里的「期」前面没有数字，不该被当题号。"""
        self.assertEqual(parser.extract_score_items("这学期八分"),
                         [(None, 8.0)])


class TestNoOpRewriteIsNotAFix(unittest.TestCase):
    """重念同一个分数（值没变）不该记成「修正」，也不该占一次撤销。"""

    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = os.path.join(self.tmp, "t.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "第一题", "第二题", "总分"])
        ws.append(["孙丽", "", "", ""])
        wb.save(self.path)
        self.state = AppState(cfg=CFG)
        self.state.load(load_sheet(self.path, CFG, backup=False))
        self.state.handle_text("孙丽")

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_same_value_is_not_counted_as_fix(self):
        self.state.handle_text("第一题八分")
        depth = len(self.state.undo_stack)
        r = self.state.handle_text("第一题八分")     # 同样的分数再念一遍
        self.assertEqual(self.state.current.scores.get(1), 8.0)
        self.assertNotIn("修正", r.message)
        self.assertEqual(len(self.state.undo_stack), depth,
                         "值没变却占了一次撤销")

    def test_different_value_still_counts_as_fix(self):
        self.state.handle_text("第一题八分")
        r = self.state.handle_text("第一题九分")
        self.assertEqual(self.state.current.scores.get(1), 9.0)
        self.assertIn("修正", r.message)


class TestNextStudentCmd(unittest.TestCase):
    """「下一个/继续」指令识别：指令词与同音姓名不能互相吞掉。"""

    def test_all_command_phrasings(self):
        for cmd in ("下一个", "下一位", "下个", "继续", "接着来", "下一名"):
            self.assertTrue(parser.is_next_student_cmd(cmd), cmd)

    def test_normal_speech_is_not_a_command(self):
        for text in ("张三", "十八分", "今天天气不错", "第一题十二分"):
            self.assertFalse(parser.is_next_student_cmd(text), text)

    def test_punctuation_stripped(self):
        self.assertTrue(parser.is_next_student_cmd("下一个。"))

    def test_homophone_name_shields_command(self):
        """「继续」和「纪旭」同音：名单里有纪旭时按姓名处理，不吞点名。"""
        self.assertFalse(parser.is_next_student_cmd("继续", ["纪旭"]))

    def test_unrelated_names_do_not_shield(self):
        self.assertTrue(parser.is_next_student_cmd("继续", ["李四", "王五"]))

    def test_short_name_never_shields(self):
        """单字名（如「旭」）发音重合面太宽，不能拿来挡指令。"""
        self.assertTrue(parser.is_next_student_cmd("继续", ["旭"]))
