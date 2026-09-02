"""Excel 读写模块：加载表格、识别结构、填写分数、计算总分、标注不一致。

使用 openpyxl 直接读写 .xlsx，保留原文件格式；首次加载自动备份。
列索引统一用 0 基（对应 openpyxl 内部会转 1 基）。
"""
from __future__ import annotations

import os
import re
import shutil
import time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------

@dataclass
class StudentRecord:
    row: int                    # Excel 行号（0 基，数据行）
    name: str
    sid: str = ""               # 学号原文（表里没有学号列时为空）
    scores: Dict[int, float] = field(default_factory=dict)   # col(0基) -> 分数
    total: Optional[float] = None
    checked: Optional[bool] = None   # True=一致 False=不一致 None=未核对
    spoken_total: Optional[float] = None   # 核对时老师念的总分，仅不一致时留着


@dataclass
class SheetModel:
    path: str
    sheet_name: str
    header: List[str] = field(default_factory=list)
    name_col: int = 0
    score_cols: List[int] = field(default_factory=list)
    total_col: int = -1
    check_col: int = -1
    header_row: int = 1            # 表头所在的 1 基行号（上方可能有标题行）
    id_col: int = -1            # 学号/座号列，-1 表示表里没有
    students: List[StudentRecord] = field(default_factory=list)

    @property
    def score_count(self) -> int:
        return len(self.score_cols)

    def next_blank_col(self, stu: StudentRecord) -> Optional[int]:
        """当前学生下一个未填分数的题号列（按表头顺序）。"""
        for col in self.score_cols:
            if col not in stu.scores:
                return col
        return None

    def calc_total(self, stu: StudentRecord) -> float:
        return round(sum(stu.scores.values()), 2)

    def filled_count(self, stu: StudentRecord) -> int:
        return sum(1 for c in self.score_cols if c in stu.scores)


# ---------------------------------------------------------------------------
# 表结构识别
# ---------------------------------------------------------------------------

def _contains_any(text: str, keys: Tuple[str, ...]) -> bool:
    return any(k in text for k in keys)


_ID_HEADERS = ("学号", "座号", "编号", "考号", "序号")


def _header_score(cells: List[str]) -> int:
    """这一行像不像表头。分数越高越像。"""
    if not any(cells):
        return -1
    score = 0
    if any(_contains_any(h, ("姓名", "名字", "学生")) for h in cells):
        score += 3
    if any(_contains_any(h, ("总分", "总成绩", "总计", "合计")) for h in cells):
        score += 2
    score += sum(1 for h in cells if _contains_any(h, ("题",)))
    return score


def find_header_row(ws, max_scan: int = 10) -> int:
    """表头所在的 1 基行号。

    老师的表常在上面留标题行（「三年级期中成绩」）或空行。写死第一行的话，
    标题行会被当成表头、真表头会被当成一个叫「姓名」的学生，整表错位，
    而且一声不响。
    """
    best_row, best = 1, -1
    upto = min(max_scan, ws.max_row or 1)
    for row in range(1, upto + 1):
        cells = [str(c.value).strip() if c.value is not None else ""
                 for c in ws[row]]
        score = _header_score(cells)
        if score > best:
            best_row, best = row, score
    return best_row


def identify_columns(header: List[str], cfg: dict) -> Tuple[int, List[int], int, int, int]:
    """根据表头与配置识别 姓名列 / 题号列 / 总分列 / 核对列 / 学号列。"""
    name_col = cfg.get("name_col", -1)
    total_col = cfg.get("total_col", -1)

    # 姓名列：配置 > 表头关键词 > 默认第一列
    if name_col < 0 or name_col >= len(header):
        name_col = 0
        for i, h in enumerate(header):
            if _contains_any(h, ("姓名", "名字", "学生", "学生姓")):
                name_col = i
                break
    name_col = max(0, min(name_col, len(header) - 1))

    # 总分列：配置 > 表头关键词 > 最后一列
    if total_col < 0 or total_col >= len(header):
        total_col = len(header) - 1
        for i, h in enumerate(header):
            if _contains_any(h, ("总分", "总成绩", "总计", "合计")):
                total_col = i
                break

    # 学号列：表头含学号/座号等关键词的列（可能在姓名列前，也可能在后）
    id_col = -1
    for i, h in enumerate(header):
        if i != name_col and _contains_any(h, _ID_HEADERS):
            id_col = i
            break

    # 题号列：姓名列之后、总分列之前，表头含"题"或纯数字，或按配置范围
    score_cols: List[int] = []
    for i, h in enumerate(header):
        if i <= name_col or i >= total_col or i == id_col:
            continue
        if _contains_any(h, ("题",)) or h.strip().isdigit():
            score_cols.append(i)
    # 兜底：姓名列与总分列之间的所有列都算题号列（学号列除外）
    if not score_cols and total_col - name_col > 1:
        score_cols = [i for i in range(name_col + 1, total_col) if i != id_col]
    # 总分列不在最右边时（如 姓名|总分|第一题|第二题），题目列全在它右边，
    # 上面按「夹在姓名与总分之间」找会一个都找不到，整张表一道题都录不进去
    if not score_cols:
        score_cols = [i for i, h in enumerate(header)
                      if i not in (name_col, total_col, id_col)
                      and (_contains_any(h, ("题",)) or h.strip().isdigit())]

    # 核对列：总分列之后含"核对/核查/检查/确认"的表头列，否则第一个空表头列
    check_col = total_col + 1
    for i in range(total_col + 1, len(header)):
        if _contains_any(header[i], ("核对", "核查", "复核", "检查", "确认")):
            check_col = i
            break
    else:
        while check_col < len(header) and header[check_col].strip():
            check_col += 1

    return name_col, score_cols, total_col, check_col, id_col


# ---------------------------------------------------------------------------
# 加载 / 保存
# ---------------------------------------------------------------------------

def load_sheet(path: str, cfg: dict, backup: bool = True) -> SheetModel:
    """加载 Excel 文件并识别结构；可选做首次备份。"""
    if not os.path.exists(path):
        raise FileNotFoundError(f"文件不存在: {path}")
    if backup:
        _make_backup(path)

    wb = load_workbook(path)
    ws = wb.active
    header_row = find_header_row(ws)
    header = [str(c.value).strip() if c.value is not None else ""
              for c in ws[header_row]]

    name_col, score_cols, total_col, check_col, id_col = identify_columns(
        header, cfg)
    # 核对列可能落在现有表头之外（原表没有这一列，保存时才写进文件）。
    # 内存里的表头必须同步补齐，否则按列索引访问会越界。
    while len(header) <= check_col:
        header.append("")
    if not header[check_col]:
        header[check_col] = cfg.get("check_header", "核对")

    model = SheetModel(
        path=path,
        sheet_name=ws.title,
        header=header,
        name_col=name_col,
        score_cols=score_cols,
        total_col=total_col,
        check_col=check_col,
        header_row=header_row,
        id_col=id_col,
    )

    for xl_row in range(header_row + 1, ws.max_row + 1):   # openpyxl 1基行号
        name_val = ws.cell(row=xl_row, column=name_col + 1).value
        if name_val is None or str(name_val).strip() == "":
            continue
        stu = StudentRecord(row=xl_row - 1, name=str(name_val).strip())  # 存0基行号
        if id_col >= 0:
            sid = ws.cell(row=xl_row, column=id_col + 1).value
            if sid is not None:
                # 学号可能被 Excel 存成数字，1001.0 要写回 1001
                stu.sid = (str(int(sid)) if isinstance(sid, float)
                           and sid.is_integer() else str(sid).strip())
        # 预填已有分数（老师手动录过一部分的情况）
        for col in score_cols:
            v = ws.cell(row=xl_row, column=col + 1).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                stu.scores[col] = float(v)
        tv = ws.cell(row=xl_row, column=total_col + 1).value
        if isinstance(tv, (int, float)) and not isinstance(tv, bool):
            stu.total = float(tv)
        elif stu.scores:
            # 总分列存的是 =SUM() 时 openpyxl 读回来的是公式字符串。
            # 照着已有分数自己算一遍，否则重开文件后总分列一片空白
            stu.total = model.calc_total(stu)
        check_val = ws.cell(row=xl_row, column=check_col + 1).value
        stu.checked = _read_checked(check_val)
        if stu.checked is False:
            stu.spoken_total = _read_spoken_total(check_val)
        model.students.append(stu)
    return model


_CHECK_OK_MARKS = frozenset({"✓", "✔", "√", "对", "ok", "yes", "一致", "正确"})
_CHECK_BAD_MARKS = frozenset({"✗", "✘", "×", "x", "错", "不一致", "错误"})
# 不一致时写的是「不一致（报62 算58）」，整体比对认不出来，按前缀认
_CHECK_BAD_PREFIX = "不一致"
_SPOKEN_RE = re.compile(r"报\s*(-?\d+(?:\.\d+)?)")


def _read_checked(value) -> Optional[bool]:
    """还原核对列的标记：一致 / 不一致 / 未核对。

    不还原就会在下次保存时把老师上一轮的核对结果清空。
    符号可能被 Excel 前后补空格，故整体比对小写去空白后的文本。
    """
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    if text in _CHECK_BAD_MARKS or text.startswith(_CHECK_BAD_PREFIX):
        return False
    if text in _CHECK_OK_MARKS:
        return True
    return None


def _read_spoken_total(value) -> Optional[float]:
    """从「不一致（报62 算58）」里取回老师当时报的那个数。"""
    if value is None:
        return None
    m = _SPOKEN_RE.search(str(value))
    return float(m.group(1)) if m else None


def check_mark_text(stu: StudentRecord) -> str:
    """核对列该写什么。

    不一致时把老师报的数一并写出来：只写「不一致」的话，老师得自己回头
    对一遍才知道差在哪；把两个数摆在一起，差多少一眼看得见。
    """
    if stu.spoken_total is None:
        return "不一致"
    expected = stu.total if stu.total is not None else 0.0
    return f"不一致（报{stu.spoken_total:g} 算{expected:g}）"


BACKUP_KEEP = 20        # 每份表格保留的备份数量上限


def _make_backup(path: str, keep: int = BACKUP_KEEP) -> Optional[str]:
    """复制一份带时间戳的备份（backups/ 目录），并清理超量的旧备份。"""
    base, ext = os.path.splitext(os.path.basename(path))
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(path)), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    ts = time.strftime("%Y%m%d-%H%M%S")
    dst = os.path.join(backup_dir, f"{base}-{ts}{ext}")
    if not os.path.exists(dst):
        shutil.copy2(path, dst)
    _prune_backups(backup_dir, base, ext, keep)
    return dst


def _prune_backups(backup_dir: str, base: str, ext: str, keep: int) -> None:
    """只保留同一份表格最近 keep 个备份，按文件名里的时间戳排序。"""
    if keep <= 0:
        return
    prefix = f"{base}-"
    try:
        names = [n for n in os.listdir(backup_dir)
                 if n.startswith(prefix) and n.endswith(ext)]
    except OSError:
        return
    for name in sorted(names, reverse=True)[keep:]:
        try:
            os.remove(os.path.join(backup_dir, name))
        except OSError:
            pass


def _center(cell) -> None:
    """把单元格设为水平垂直居中，其余对齐属性（换行、缩进等）原样保留。"""
    a = cell.alignment
    if a.horizontal == "center" and a.vertical == "center":
        return
    cell.alignment = Alignment(
        horizontal="center", vertical="center",
        text_rotation=a.text_rotation, wrap_text=a.wrap_text,
        shrink_to_fit=a.shrink_to_fit, indent=a.indent)


_RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_RED_FONT = Font(color="FF9C0006", bold=True)
_GREEN_FONT = Font(color="FF006100")


def save_sheet(model: SheetModel, cfg: dict, write_formula: bool = True) -> str:
    """把内存数据写回 Excel 文件，返回保存路径。

    - 分数：写入题号列
    - 总分：只写已录入分数的行（一分没录的学生保持原样，避免整列被写成 0）；
      写公式 =SUM(区间) 可由 Excel/WPS 自动重算，软件内对比用内存计算结果
    - 核对列：不一致 -> "不一致"红底红字；一致 -> "✓"绿色（可配置不写）
    """
    wb = load_workbook(model.path)   # 保留原有格式
    ws = wb[model.sheet_name]

    # 确保核对列表头存在
    check_header = ws.cell(row=model.header_row, column=model.check_col + 1)
    if check_header.value is None or str(check_header.value).strip() == "":
        check_header.value = cfg.get("check_header", "核对")

    has_score_cols = bool(model.score_cols)
    if has_score_cols:
        first_score, last_score = model.score_cols[0], model.score_cols[-1]
    for stu in model.students:
        xl_row = stu.row + 1
        # 1) 分数：内存里删掉的分数（撤销、清空该生）必须同步清出表格，
        #    否则文件里的 =SUM() 仍会把旧分数算进总分。
        #    只清数值格，老师手写的「缺考」这类文字保留。
        cleared_any = False
        for col in model.score_cols:
            cell = ws.cell(row=xl_row, column=col + 1)
            if col in stu.scores:
                cell.value = float(stu.scores[col])
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.value = None
                cleared_any = True
        for col, val in stu.scores.items():
            if col not in model.score_cols:   # 配置指定了题号列范围之外的分数
                ws.cell(row=xl_row, column=col + 1, value=float(val))
        # 2) 总分：没录过分的行不碰，否则 =SUM(空白) 会把全班总分写成 0
        total_cell = ws.cell(row=xl_row, column=model.total_col + 1)
        if stu.scores:
            if write_formula and has_score_cols:
                letter_a = get_column_letter(first_score + 1)
                letter_b = get_column_letter(last_score + 1)
                total_cell.value = f"=SUM({letter_a}{xl_row}:{letter_b}{xl_row})"
            elif stu.total is not None:
                total_cell.value = stu.total
        elif cleared_any:
            total_cell.value = None   # 分数被清空，留着公式会算成 0
        # 3) 核对列
        cell = ws.cell(row=xl_row, column=model.check_col + 1)
        if stu.checked is None:
            # 只清掉本程序写过的标记，老师自己填的批注保留
            if _read_checked(cell.value) is not None:
                cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()
        elif stu.checked:
            if cfg.get("write_checked", True):
                cell.value = "✓"
                cell.font = _GREEN_FONT
                cell.fill = PatternFill()
            else:
                cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()
        else:
            cell.value = check_mark_text(stu)
            cell.fill = _RED_FILL
            cell.font = _RED_FONT

    _center_used_range(ws, model)
    wb.save(model.path)
    return model.path


def _center_used_range(ws, model: SheetModel) -> None:
    """表头行与各学生行一律居中，跟软件里看到的表格保持一致。"""
    last_col = max(model.check_col, len(model.header) - 1)
    for col in range(last_col + 1):
        _center(ws.cell(row=model.header_row, column=col + 1))
    for stu in model.students:
        for col in range(last_col + 1):
            _center(ws.cell(row=stu.row + 1, column=col + 1))
    # 核对列要放下「不一致（报90 算100）」，默认宽度会显示成 ####
    letter = get_column_letter(model.check_col + 1)
    need = max([_display_width(
                    ws.cell(row=model.header_row,
                            column=model.check_col + 1).value)]
               + [_display_width(check_mark_text(s) if s.checked is False
                                 else "✓" if s.checked else "")
                  for s in model.students]) + 2
    if (ws.column_dimensions[letter].width or 0) < need:
        ws.column_dimensions[letter].width = need


def _display_width(value) -> int:
    """Excel 列宽按半角字符数算，汉字与全角标点各占两格。"""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in str(value or ""))