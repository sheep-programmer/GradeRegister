#!/usr/bin/env python3
"""生成一张示例成绩表 sample.xlsx，方便体验系统。

表结构：姓名 | 第一题 | 第二题 | 第三题 | 第四题 | 总分
其中包含两个「张三」用于演示重名确认弹窗。
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER = ["姓名", "第一题", "第二题", "第三题", "第四题", "总分"]
ROWS = [
    ["张三", "", "", "", ""],
    ["李四", "", "", "", ""],
    ["王五", "", "", "", ""],
    ["张三", "", "", "", ""],
    ["赵六", "", "", "", ""],
    ["钱七", "", "", "", ""],
]


def main() -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩表"

    fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    bold = Font(bold=True)
    for c, name in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(ROWS, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)

    for col, width in zip("ABCDEF", (10, 10, 10, 10, 10, 10)):
        ws.column_dimensions[col].width = width

    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample.xlsx")
    wb.save(path)
    print(f"示例表格已生成: {path}")
    print("提示：表里有两位「张三」，念『张三』时会弹窗让您确认是哪一行，"
          "正好演示重名选择功能。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())